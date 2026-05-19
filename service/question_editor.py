# openpyxl
import openpyxl as op
from openpyxl import Workbook
from openpyxl.styles import Protection, Alignment
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.styles import PatternFill, Font

# project
from database.connector import SQLite3Connector as sqlite3_connector
from service.question import QuestionSet
from config_data.config import config
from utils.misc import replace_space

# misc
import re
import datetime




class Excel():
    header = ['q_id', 'Question', 'r_num']

    def download_questions(account_id: int) -> str:

        # Create a new workbook
        wb = Workbook()

        # Get the client name
        account_name = replace_space(sqlite3_connector.get_account_name(account_id))

        # Get all sets linked to this client
        sets = sqlite3_connector.get_sets_by_account(account_id)

        # Header for the users tab
        # We immediately record set_id and set_name here
        sets_header = []

        for set_ in sets:
            # Record "set name (set number)" in the header
            sets_header.append(f'{set_[1]} ({set_[0]})')
            Excel._create_questions_worksheet(wb, set_)

        # Remove the first sheet (created automatically when the workbook was created)
        wb.remove(wb.worksheets[0])

        Excel._create_user_worksheet(account_id, wb, sets_header)
        
        # Save the file
        file_name = f'{config.export_file_storage}update_{account_id}_{account_name}_{datetime.datetime.now().strftime("%d-%m-%Y_%H-%M-%S")}.xlsx'
        wb.save(file_name)

        return file_name

    def _create_user_worksheet(account_id, wb, sets_header):
        
        # Create a new users tab
        user_ws = wb.create_sheet('Users')

        # Add set columns to the main header
        header = ['user_id', 'Telegram', 'Full Name', 'Active', *sets_header]

        user_ws.append(header)

        # Iterate over each user
        for user in sqlite3_connector.get_all_users_by_account_id(account_id):
            # Array to be added as a row in the table for each user
            append_row = []
            append_row.append(user[0])  # user_id
            append_row.append(user[1]) # telegram username
            append_row.append(user[2]) # real name

            # It is better to use words for user activity rather than a 1
            append_row.append('Active' if user[3] else 'Inactive')

            # Extend the array to the number of sets
            append_row.extend([''] * len(sets_header))

            # Iterate over each of the user's sets
            for user_sets in sqlite3_connector.get_users_set_list(user_id = user[0]):
                index = sets_header.index(f'{user_sets[1]} ({user_sets[0]})')
                append_row[4+index] = user_sets[2]

            # Add the user record
            user_ws.append(append_row)

    def _create_questions_worksheet(wb, set_):

        # Create a new sheet in the workbook named "set name (set number)"
        ws = wb.create_sheet(f'{set_[1]} ({set_[0]})')

        # Variable to count the maximum number of answers. Needed to correctly form the table header.
        max_answers = 0

        # Get the array of question objects
        question_set = QuestionSet(set_[0])

        # Add each question with its answers to the table,
        # also track the maximum number of answers
        for question in question_set.questions:
            ws.append(question._form_row())
            max_answers = max(max_answers, len(question.answers))

        # Set the table header
        header = ['q_id', 'Question', 'Number of correct answers']

        # Append answer columns based on the maximum number of answers
        for i in range(1, max_answers+1):
            header.append('a_id')
            header.append(f'Answer - {i}')
            
        # Add the header row
        if len(question_set.questions) == 0:
            ws.append(header)
        else:
            ws.insert_rows(1)

        # Format the header
        for row in ws.iter_rows(min_row = 1, max_col = ws.max_column, max_row = 1):
            for index, cell in enumerate(row):
                
                # Set column header titles
                cell.value = header[index]

                # Bold and underlined text, size 12
                cell.font = Font(bold = True, 
                                     underline = 'single',
                                     size = 12)
                    
                # Grey background
                cell.fill = PatternFill(start_color='D3D3D3', 
                                            end_color='D3D3D3', 
                                            fill_type='solid')

                # Center alignment and text wrapping
                cell.alignment = Alignment(wrap_text = True, 
                                              vertical = 'center', 
                                              horizontal = 'center')

                # If the column stores an id, hide it
                if cell.value.find('id') >= 0:
                    ws.column_dimensions[op.utils.cell.get_column_letter(index+1)].hidden = True
                        
                # All other columns have fixed widths
                else:
                    
                    # For the column with the number of correct answers
                    if index+1 == 3:
                        ws.column_dimensions[op.utils.cell.get_column_letter(index+1)].width = 15
                        
                    # For question columns
                    else:
                        ws.column_dimensions[op.utils.cell.get_column_letter(index+1)].width = 300 / 7
                        
                    ws.column_dimensions[op.utils.cell.get_column_letter(index+1)].auto_size = False

                    # Unlock cells that the user can edit
                    for unlock_cells in ws[op.utils.cell.get_column_letter(index+1)]:

                        # Unlock everything below row 1
                        if unlock_cells.row > 1:
                            unlock_cells.protection = Protection(locked = False)
                            
                            # Center the column with the number of correct answers
                            if unlock_cells.column == 3:
                                unlock_cells.alignment = Alignment(wrap_text = True, 
                                                                       vertical = 'center', 
                                                                       horizontal = 'center')
                            # All others just vertically center-aligned
                            else:
                                unlock_cells.alignment = Alignment(wrap_text = True, 
                                                                       vertical = 'top')
                                    
                    ws.column_dimensions[op.utils.cell.get_column_letter(index+1)].protection = Protection(locked = False)

        # Set protection format (everything locked)
        protection = SheetProtection(sheet=True, objects=True, scenarios=True,
                                 formatCells=True, formatColumns=True, formatRows=True,
                                 insertColumns=True, insertRows=True, insertHyperlinks=True,
                                 deleteColumns=True, deleteRows=True, selectLockedCells=True,
                                 sort=True, autoFilter=True, pivotTables=True)
    
        # Allow resizing of columns
        protection.formatRows = False
            
        # Apply protection to the sheet
        ws.protection = protection

    def upload_questions(account_id: int, file_path: str) -> None:

        # Open the workbook
        wb = op.load_workbook(file_path)


        users_ws = wb['Users']

        # Check if the users sheet exists
        if users_ws != None:
            Excel._update_users(users_ws, account_id)

        # Iterate over each sheet
        for sheet in wb.worksheets:

            if sheet.title == 'Users':
                continue

            set_id = 0

            # Find N in "Name (N)"
            match = re.search(r"\((\d+)\)" , sheet.title)

            # If found, that is our set id
            if match:
                set_id = match.group(1)

            # If not found, we need to create a set in the database
            else:
                set_id = sqlite3_connector.create_set(sheet.title)

            set_title = sheet.title

            # Delete the header
            sheet.delete_rows(0)

            # Iterate over each row (using cell values only)
            for row in sheet.iter_rows(values_only = True):

                if not any(row):
                    continue

                question_id = row[0]
                question_text = row[1]
                right_answers_count = int(row[2])

                # If the question has an id
                # (i.e. we are editing an existing question)
                if type(question_id) == int:
                    
                    # If there is an id but no text,
                    # the user deleted the question
                    if question_text is None or len(question_text.replace(' ','')) == 0:
                        sqlite3_connector.remove_question(question_id)
                        continue # No point continuing since we already deleted all answers
                    
                    # If the question text exists, update it in the database,
                    # since the user may have edited it
                    else:
                        sqlite3_connector.update_question(question_id, question_text)
                
                # If the question has no id,
                # it is a new question
                # and must be added to the database
                else:
                    question_id = sqlite3_connector.add_question(set_id, account_id, question_text)
                

                # Iterate over the answers
                # with step 2 (since each answer has a text and an id)
                for i in range(3, len(row), 2):
                    answer_id = row[i]
                    answer_text = row[i+1]

                    # This line is temporary, in case the user sets
                    # the number of correct answers to more than 1.
                    # In the future the variable should simply take the value without conditions.
                    right = 1 if right_answers_count > 0 else 0

                    if type(answer_id) == int:
                        sqlite3_connector.update_answer(answer_id, answer_text, right)
                    elif answer_text != None:
                        sqlite3_connector.add_answer(question_id, answer_text, right)

                    # Decrease the number of correct answers
                    right_answers_count -= 1

    def _update_users(users_ws, account_id):

        # Get the header
        header = [cell.value for cell in next(users_ws.iter_rows(min_row=1, max_row=1))]

        # Rename set titles, leaving only their ids in the cells
        for set_ in header[4:]:
            header[header.index(set_)] = int(re.search(r"\((\d+)\)" , set_).group(1))

        header = header[4:]

        # Delete the header row in Excel
        users_ws.delete_rows(0)

        # Iterate over all users
        for row in users_ws.iter_rows(values_only = True):

            # Assign values
            user_id = row[0]                            # user id
            telegram_username = row[1]                  # telegram login
            real_name = row[2]                          # full name
            active = 0 if row[3] == 'Inactive' else 1  # active or not (1, 0)

            # If there is no user_id in the cell, this is a new user
            if user_id == None:
                user_id = sqlite3_connector.add_user(telegram_username = telegram_username,
                                                     real_name = real_name,
                                                     account_id = account_id,
                                                     active = active)

            # If there is, update the existing user
            else:
                sqlite3_connector.update_user(user_id = user_id,
                                              telegram_username = telegram_username,
                                              real_name = real_name,
                                              active = active)
            
            # Iterate over the sets
            for i, qty in enumerate(row[4:]):
                sqlite3_connector.set_set_qty(user_id = user_id, set_id = header[i], qty = qty)
